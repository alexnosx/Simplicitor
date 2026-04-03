# simplicitor/app/generators/excel_generator.py
# Phase 3: Excel spreadsheet generator
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Generates .xlsx files from parsed LLM output (Phase 3)."""

    def generate(self, parsed: dict, output_path: Path) -> Path:
        """Write an Excel spreadsheet from the parsed LLM response structure.

        Args:
            parsed: Validated dict from parse_excel_response(). Expected keys:
                    ``sheet_name`` (str), ``headers`` (list[str]),
                    ``rows`` (list[list]), and ``formulas`` (list[dict]).
            output_path: Full file path (including filename) to write.

        Returns:
            output_path as a Path.

        Raises:
            OSError: On disk write failure.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = parsed.get("sheet_name", "Sheet1")

        headers = parsed.get("headers", [])
        rows = parsed.get("rows", [])
        formulas = parsed.get("formulas", [])

        # Write headers in row 1 with bold font
        bold_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold_font

        # Write data rows starting at row 2
        for row_idx, row_data in enumerate(rows, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=self._coerce_value(value))

        # Apply formulas
        for formula_entry in formulas:
            cell_ref = formula_entry.get("cell", "")
            formula = formula_entry.get("formula", "")
            if cell_ref and formula:
                ws[cell_ref] = formula

        # Auto-fit column widths
        self._auto_fit_columns(ws)

        wb.save(str(output_path))
        return output_path

    def _coerce_value(self, value: object) -> object:
        """Try to convert a string value to int or float; return as-is otherwise."""
        if not isinstance(value, str):
            return value
        # Attempt integer conversion first
        try:
            return int(value)
        except (ValueError, TypeError):
            pass
        # Attempt float conversion
        try:
            return float(value)
        except (ValueError, TypeError):
            pass
        return value

    def _auto_fit_columns(self, ws) -> None:
        """Set each column's width to max content length + 2 for readability."""
        column_widths: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    col_letter = get_column_letter(cell.column)
                    content_len = len(str(cell.value))
                    if col_letter not in column_widths or content_len > column_widths[col_letter]:
                        column_widths[col_letter] = content_len

        for col_letter, max_len in column_widths.items():
            ws.column_dimensions[col_letter].width = max_len + 2
